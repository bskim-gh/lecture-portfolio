"""
Ch09 - Python 웹 크롤링
=======================

주요 내용:
- 웹 크롤링(Web Crawling) 개념
- requests 모듈을 이용한 웹 페이지 요청
- BeautifulSoup을 이용한 HTML 파싱
- CSS Selector를 이용한 요소 선택
- select_one(), select() 메서드
- 네이버 뉴스 크롤링
- 인터파크 항공권 크롤링
- Selenium을 이용한 동적 크롤링
- 가상 브라우저(ChromeDriver) 사용
- 웹 요소 찾기: find_element_by_css_selector(), find_element_by_id()
- 웹 요소 조작: click(), send_keys()
- 크롤링 데이터를 Excel 파일로 저장
- 날씨 정보 크롤링 및 DB 저장
"""

import requests as req
from bs4 import BeautifulSoup as bs

# ====================================
# 1. requests를 이용한 웹 요청
# ====================================

def naver_request():
    # 네이버 요청
    resp = req.get('https://www.naver.com')
    print(resp.text[:200])  # 처음 200자만 출력

    # 파싱(HTML문서에서 특정 데이터 추출)
    dom = bs(resp.text, 'html.parser')
    # tag_a = dom.select_one('#yna_rolling > div:nth-child(8) > a')
    # print(tag_a.text)


# ====================================
# 2. CSS Selector를 이용한 요소 선택
# ====================================

def css_selector_example():
    resp = req.get('http://chhak.kr/py/test1.html')
    resp.encoding = 'utf-8'
    print(resp.text)

    # 파싱
    dom = bs(resp.text, 'html.parser')

    tag_tit = dom.html.body.h1
    tag_txt = dom.select_one('#txt')
    tag_li1 = dom.select_one('ul > li:nth-child(1)')
    tag_li2 = dom.select_one('ul > li:nth-child(2)')
    tag_li3 = dom.select_one('ul > li:nth-child(3)')
    tag_li4 = dom.select_one('ul > li:nth-child(4)')
    tag_li5 = dom.select_one('ul > li:last-child')
    tag_lis = dom.select('ul > li')

    print(tag_tit.text)
    print(tag_txt.text)
    print(tag_li1.text)
    print(tag_li2.text)
    print(tag_li3.text)
    print(tag_li4.text)
    print(tag_li5.text)
    print(tag_lis)

    for li in tag_lis:
        print(li.text)


# ====================================
# 3. 인터파크 항공권 크롤링
# ====================================

def interpark_crawling():
    from openpyxl import Workbook

    day = 10

    # Excel 파일생성
    workbook = Workbook()
    sheet = workbook.active

    while True:
        print('day :', day)
        url = 'http://domair.interpark.com/dom/main.do?trip=OW&dep=GMP&arr=PUS&dep2=PUS&arr2=GMP&depdate=202103%d&retdate=20210310&adt=1&chd=0&inf=0&mbn=tour_main&mln=search_domair0#anchor-list' % day
        resp = req.get(url, headers={'User-Agent': 'Mozilla/5.0'})

        dom = bs(resp.text, 'html.parser')
        hl = dom.select('#availDepResult > table > tbody > tr > td')

        for a in hl:
            print(a.text.strip())
            sheet.append([a.text.strip()])

        day += 1

        if day > 20:
            break

    # Excel 파일 저장
    workbook.save('C:/Users/user/Desktop/Fare.xlsx')


# ====================================
# 4. 네이버 뉴스 크롤링
# ====================================

def naver_news_crawling():
    from openpyxl import Workbook

    page = 1

    # Excel 파일생성
    workbook = Workbook()
    sheet = workbook.active

    while True:
        print('page :', page)
        url = 'https://news.naver.com/main/list.nhn?mode=LSD&mid=sec&sid1=001&date=20210308&page=' + str(page)
        resp = req.get(url, headers={'User-Agent': 'Mozilla/5.0'})

        dom = bs(resp.text, 'html.parser')
        hl = dom.select('#main_content > div.list_body.newsflash_body > ul > li > dl > dt:nth-child(2) > a')

        for a in hl:
            print(a.text.strip())
            sheet.append([a.text.strip()])

        page += 1

        if page > 10:
            break

    # Excel 파일 저장
    workbook.save('C:/Users/user/Desktop/News.xlsx')


# ====================================
# 5. Selenium을 이용한 동적 크롤링
# ====================================

def selenium_naver_login():
    from selenium import webdriver

    # 가상브라우저 실행
    browser = webdriver.Chrome('chromedriver.exe')

    # 네이버 이동
    browser.get('http://naver.com')

    # 네이버 로그인 버튼 클릭
    btn_login = browser.find_element_by_css_selector('#account > a')
    btn_login.click()

    # 네이버 아이디, 비밀번호 입력
    input_id = browser.find_element_by_css_selector('#id')
    input_pw = browser.find_element_by_id('pw')
    btn_login2 = browser.find_element_by_css_selector('#log\\.login')

    input_id.send_keys('')
    input_pw.send_keys('')
    btn_login2.click()


# ====================================
# 6. Selenium을 이용한 네이버 뉴스 수집
# ====================================

def selenium_news_crawling():
    from selenium import webdriver
    from openpyxl import Workbook

    browser = webdriver.Chrome('chromedriver.exe')

    # 네이버 이동
    browser.get('http://naver.com')

    # 네이버 로그인 버튼 클릭
    btn_news = browser.find_element_by_css_selector('#NM_FAVORITE > div.group_nav > ul.list_nav.NM_FAVORITE_LIST > li:nth-child(2) > a')
    btn_news.click()

    btn_breaking = browser.find_element_by_css_selector('#lnb > ul > li:nth-child(2) > a > span')
    btn_breaking.click()

    page = 2

    # Excel 파일생성
    workbook = Workbook()
    sheet = workbook.active

    while True:
        print('----------------------------------------------')
        print('page :', page-1)
        tags_title = browser.find_elements_by_css_selector('#main_content > div.list_body.newsflash_body > ul > li > dl > dt:nth-child(2) > a')

        for tit in tags_title:
            print(tit.text)
            sheet.append([tit.text.strip()])

        page_btn = "#main_content > div.paging > a:nth-child(%d)" % page
        btn_page = browser.find_element_by_css_selector(page_btn)
        btn_page.click()

        page += 1

        if page > 10:
            print('----------------------------------------------')
            break

    # Excel 파일 저장
    workbook.save('C:/Users/user/Desktop/News2.xlsx')

    # 브라우저 종료
    browser.quit()


# ====================================
# 7. 네이버 뉴스 수집 (날짜 포함)
# ====================================

def selenium_news_with_date():
    from selenium import webdriver
    from openpyxl import Workbook
    from datetime import datetime

    # 가상브라우저 실행
    browser = webdriver.Chrome('chromedriver.exe')

    # 네이버 이동
    browser.get('https://naver.com')

    # 뉴스 클릭
    btn_news = browser.find_element_by_css_selector('#NM_FAVORITE > div.group_nav > ul.list_nav.NM_FAVORITE_LIST > li:nth-child(2) > a')
    btn_news.click()

    # 속보 클릭
    btn_sokbo = browser.find_element_by_css_selector('#lnb > ul > li:nth-child(2) > a')
    btn_sokbo.click()

    i = 0

    # Excel 파일생성
    workbook = Workbook()
    sheet = workbook.active

    while True:
        # 뉴스 리스트 제목 선택
        tags_dl = browser.find_elements_by_css_selector('#main_content > div.list_body.newsflash_body > ul > li > dl')

        # 뉴스 제목 출력
        for dl in tags_dl:
            try:
                title   = dl.find_element_by_css_selector('dt:nth-child(2) > a')
                writer  = dl.find_element_by_css_selector('dd > span.writing')
                now     = "{:%Y-%m-%d}".format(datetime.now())

                print('title :', title.text.strip())
                print('writer :', writer.text)
                sheet.append([writer.text, title.text.strip(), now])
            except:
                print('예외발생')

        # 다음 페이지 번호 클릭
        tags_paging = browser.find_elements_by_css_selector('#main_content > div.paging > a')
        tags_paging[i].click()

        i += 1

        # 10번까지 제목 출력 후 종료
        if i > 8:
            break

    # 브라우저 종료
    browser.quit()

    # Excel 파일 저장
    workbook.save('C:/Users/user/Desktop/News2.xlsx')


# ====================================
# 8. 날씨 정보 크롤링 및 DB 저장
# ====================================

def weather_crawling_to_db():
    from selenium import webdriver
    import pymysql

    # 가상브라우저 실행
    browser = webdriver.Chrome('chromedriver.exe')
    browser.get('https://www.weather.go.kr/w/obs-climate/land/city-obs.do')

    trs = browser.find_elements_by_css_selector('#weather_table > tbody > tr')

    # 데이터베이스 접속
    conn = pymysql.connect(host='공용DB localhost',
                           user='bskim02',
                           password='1234',
                           db='bskim02_db',
                           charset='utf8')

    cursor = conn.cursor()

    for tr in trs:
        city     = tr.find_element_by_css_selector('td > a').text
        visible  = tr.find_element_by_css_selector('td:nth-child(3)').text
        temp     = tr.find_element_by_css_selector('td:nth-child(6)').text
        dew      = tr.find_element_by_css_selector('td:nth-child(7)').text
        feel     = tr.find_element_by_css_selector('td:nth-child(8)').text
        humidity = tr.find_element_by_css_selector('td:nth-child(11)').text
        wind_dir = tr.find_element_by_css_selector('td:nth-child(12)').text
        wind_spd = tr.find_element_by_css_selector('td:nth-child(13)').text
        sea      = tr.find_element_by_css_selector('td:nth-child(14)').text

        # SQL 실행
        sql = "INSERT INTO `Weather` SET "
        sql += "`city`='"+city+"',"
        sql += "`visible`='"+visible+"',"
        sql += "`temp`='"+temp+"',"
        sql += "`dew`='"+dew+"',"
        sql += "`feel`='"+feel+"',"
        sql += "`humidity`='"+humidity+"',"
        sql += "`wind_dir`='"+wind_dir+"',"
        sql += "`wind_spd`='"+wind_spd+"',"
        sql += "`sea`='"+sea+"',"
        sql += "`rdate`= NOW();"

        cursor.execute(sql)
        conn.commit()

        print('수집 중..')

    # 데이터베이스 종료
    conn.close()

    # 브라우저 종료
    browser.quit()

    print('수집 완료!')


if __name__ == '__main__':
    # 기본 요청 예제 실행
    naver_request()
    
    print('Ch09 웹 크롤링 학습 완료')
    # 실행 예제: css_selector_example(), naver_news_crawling() 등
